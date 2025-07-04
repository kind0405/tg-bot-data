import os
import io
import json
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import docx
import openpyxl
import PyPDF2

# --- НАСТРОЙКИ ---
SERVICE_ACCOUNT_FILE = 'credentials.json'  # путь к твоему service account json
FOLDER_ID = '1SDBfV-2Zk7lriKUsgRSS6wWnyC2O7ZX0'  # твоя папка на Google Диске
SCOPES = ['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/documents', 'https://www.googleapis.com/auth/spreadsheets.readonly']

# --- АВТОРИЗАЦИЯ ---
credentials = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES)
drive_service = build('drive', 'v3', credentials=credentials)
docs_service = build('docs', 'v1', credentials=credentials)
sheets_service = build('sheets', 'v4', credentials=credentials)

def chunk_text(text, chunk_size=500):
    chunks = []
    start = 0
    length = len(text)
    while start < length:
        end = start + chunk_size
        if end >= length:
            chunks.append(text[start:].strip())
            break
        else:
            space_pos = text.rfind(' ', start, end)
            if space_pos == -1 or space_pos <= start:
                space_pos = end
            chunks.append(text[start:space_pos].strip())
            start = space_pos
    return chunks

def extract_text_from_google_doc(doc_id):
    try:
        doc = docs_service.documents().get(documentId=doc_id).execute()
        content = doc.get('body').get('content')
        text = ''
        for c in content:
            paragraph = c.get('paragraph')
            if paragraph:
                elements = paragraph.get('elements')
                for e in elements:
                    text_run = e.get('textRun')
                    if text_run:
                        text += text_run.get('content')
        return text.strip()
    except Exception as e:
        print(f"Ошибка при чтении Google Документа {doc_id}: {e}")
        return ''

def extract_text_from_google_sheet(sheet_id):
    try:
        # Получаем все листы
        spreadsheet = sheets_service.spreadsheets().get(spreadsheetId=sheet_id).execute()
        sheets = spreadsheet.get('sheets')
        all_text = ''
        for sheet in sheets:
            title = sheet['properties']['title']
            # Читаем весь диапазон листа
            result = sheets_service.spreadsheets().values().get(
                spreadsheetId=sheet_id,
                range=title
            ).execute()
            values = result.get('values', [])
            # Преобразуем в текст, разделяем табуляцией
            for row in values:
                line = '\t'.join(row)
                all_text += line + '\n'
        return all_text.strip()
    except Exception as e:
        print(f"Ошибка при чтении Google Таблицы {sheet_id}: {e}")
        return ''

def extract_text_from_pdf(file_stream):
    try:
        reader = PyPDF2.PdfReader(file_stream)
        text = ''
        for page in reader.pages:
            text += page.extract_text() + '\n'
        return text.strip()
    except Exception as e:
        print(f"Ошибка при чтении PDF: {e}")
        return ''

def extract_text_from_docx(file_stream):
    try:
        doc = docx.Document(file_stream)
        fullText = []
        for para in doc.paragraphs:
            fullText.append(para.text)
        return '\n'.join(fullText).strip()
    except Exception as e:
        print(f"Ошибка при чтении DOCX: {e}")
        return ''

def extract_text_from_xlsx(file_stream):
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        all_text = ''
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = [str(cell) if cell is not None else '' for cell in row]
                all_text += '\t'.join(row_text) + '\n'
        return all_text.strip()
    except Exception as e:
        print(f"Ошибка при чтении XLSX: {e}")
        return ''

def download_file(file_id, mime_type):
    """Скачиваем файл с Google Диска в память"""
    request = drive_service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
    fh.seek(0)
    return fh

def get_files_from_folder(folder_id):
    query = f"'{folder_id}' in parents and trashed = false"
    files = []
    page_token = None
    while True:
        response = drive_service.files().list(
            q=query,
            spaces='drive',
            fields='nextPageToken, files(id, name, mimeType)',
            pageToken=page_token
        ).execute()
        files.extend(response.get('files', []))
        page_token = response.get('nextPageToken', None)
        if page_token is None:
            break
    return files

def main():
    print("Получаем список файлов из папки...")
    files = get_files_from_folder(FOLDER_ID)
    data = []

    for f in files:
        file_id = f['id']
        name = f['name']
        mime = f['mimeType']
        print(f"Обрабатываем файл: {name} ({mime})")

        text = ''
        try:
            if mime == 'application/vnd.google-apps.document':
                text = extract_text_from_google_doc(file_id)
            elif mime == 'application/vnd.google-apps.spreadsheet':
                text = extract_text_from_google_sheet(file_id)
            elif mime == 'application/vnd.google-apps.presentation':
                # Можно добавить, если нужно, обработку презентаций (сейчас пропускаем)
                print("Обработка Google Презентаций не реализована, пропускаем.")
                continue
            elif mime == 'application/pdf':
                fh = download_file(file_id, mime)
                text = extract_text_from_pdf(fh)
            elif mime in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document']:
                fh = download_file(file_id, mime)
                text = extract_text_from_docx(fh)
            elif mime in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                fh = download_file(file_id, mime)
                text = extract_text_from_xlsx(fh)
            else:
                print(f"Неизвестный тип файла {mime}, пропускаем.")
                continue
        except Exception as e:
            print(f"Ошибка при обработке файла {name}: {e}")
            continue

        if not text.strip():
            print(f"Пустой текст в файле {name}, пропускаем.")
            continue

        data.append({
            "id": file_id,
            "title": name,
            "text_chunks": chunk_text(text, chunk_size=500),
            "url": f"https://drive.google.com/file/d/{file_id}/view?usp=sharing",
            "embedding": []
        })

    # Сохраняем в JSON
    with open('data.json', 'w', encoding='utf-8') as fjson:
        json.dump(data, fjson, ensure_ascii=False, indent=2)

    print("Готово! Файл data.json создан.")

if __name__ == '__main__':
    main()